"""
Выгрузка отчёта для Финнадзора КР (XLSX).

Формат листов повторяет Приложения 4/о («Продажа») и 5/о («Покупка»)
ежемесячного оперативного отчёта операторов обмена виртуальных активов.
Одна строка = одна транзакция (части сделки «тест + остаток» идут
отдельными строками с рублями, рассчитанными по курсу сделки).
"""
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from services.doc_generator import computed_transactions, _to_float

MONTHS_RU_NOM = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

HEADERS = [
    "№",
    "Дата сделки",
    "Наименование и ИНН (для юридического лица) или Ф.И.О. и ID (для физического лица)",
    "Резидент/нерезидент Кыргызской Республики",
    "Наименование виртуального актива",
    "Курс обмена ВА на денежные средства",
    "Объём в денежных средствах (в иностранной валюте)",
    "Объём в денежных средствах (в сомах)",
    "Метод расчета (наличный/безналичный)",
    "Адрес электронного кошелька оператора",
    "Адрес электронного кошелька клиента",
    "Идентификация клиента KYC (да/нет)",
]

_thin = Side(style="thin")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_font = Font(name="Times New Roman", size=10)
_bold = Font(name="Times New Roman", size=10, bold=True)
_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

COL_WIDTHS = [5, 12, 38, 16, 14, 12, 18, 18, 15, 36, 36, 14]


def _deal_rows(deal: dict, kgs_rate: float) -> list:
    """Одна сделка → строки-транзакции для листа отчёта."""
    deal_like = {
        "va_amount": deal.get("va_amount", 0),
        "fiat_amount": deal.get("fiat_amount", 0),
        "tx_hash": "",
        "transactions": json.loads(deal.get("transactions") or "[]"),
    }
    txs = computed_transactions(deal_like)

    try:
        deal_date = datetime.strptime(deal.get("deal_date_iso", ""), "%Y-%m-%d").date()
    except ValueError:
        deal_date = deal.get("deal_date", "")

    client = f"{deal.get('client_name','')} {deal.get('client_inn','')}".strip()
    rate = _to_float(deal.get("exchange_rate"))

    rows = []
    for t in txs:
        fiat = float(t.get("fiat_amount") or 0)
        rows.append([
            deal_date,
            client,
            deal.get("client_resident") or "Резидент",
            deal.get("va_type", ""),
            rate,
            fiat,
            round(fiat * kgs_rate, 2) if kgs_rate else "",
            "Безналичный",
            deal.get("operator_wallet", ""),
            deal.get("client_wallet", ""),
            "Да",
        ])
    return rows


def _fill_sheet(ws, title: str, annex: str, deals: list, kgs_rate: float):
    ws.cell(row=1, column=len(HEADERS) - 1, value=annex).font = _bold
    ws.cell(row=2, column=1, value=title).font = _bold
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    for col, header in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=header)
        c.font = _bold
        c.alignment = _wrap
        c.border = _border

    row_idx = 4
    n = 0
    total_fiat = total_som = 0.0
    for deal in deals:
        for values in _deal_rows(deal, kgs_rate):
            n += 1
            ws.cell(row=row_idx, column=1, value=n)
            for col, v in enumerate(values, 2):
                ws.cell(row=row_idx, column=col, value=v)
            for col in range(1, len(HEADERS) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.font = _font
                c.border = _border
                if col == 2 and not isinstance(values[0], str):
                    c.number_format = "DD.MM.YYYY"
                if col in (7, 8):
                    c.number_format = "#,##0.00"
            total_fiat += float(values[5] or 0)
            total_som += float(values[6] or 0)
            row_idx += 1

    # Итого
    ws.cell(row=row_idx, column=1, value="Итого").font = _bold
    c_fiat = ws.cell(row=row_idx, column=7, value=round(total_fiat, 2))
    c_som = ws.cell(row=row_idx, column=8, value=round(total_som, 2))
    for c in (c_fiat, c_som):
        c.font = _bold
        c.number_format = "#,##0.00"
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=row_idx, column=col).border = _border

    for i, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = width
    ws.freeze_panes = "A4"


def build_fn_report(path: str, month: str, deals: list, kgs_rate: float) -> str:
    """
    month  — 'YYYY-MM'
    deals  — строки таблицы deals за месяц
    kgs_rate — курс RUB→KGS для колонки «в сомах» (0 = не заполнять)
    """
    sell = [d for d in deals if d.get("act_type") == "sell"]
    buy = [d for d in deals if d.get("act_type") == "buy"]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Продажа"
    _fill_sheet(
        ws1, "Отчет по продаже виртуальных активов клиентам",
        "Приложение 4/о", sell, kgs_rate,
    )
    ws2 = wb.create_sheet("Покупка")
    _fill_sheet(
        ws2, "Отчет по покупке виртуальных активов у клиентов",
        "Приложение 5/о", buy, kgs_rate,
    )
    wb.save(path)
    return path


def month_label(month: str) -> str:
    """'2026-06' → 'Июнь 2026'"""
    try:
        y, m = month.split("-")
        return f"{MONTHS_RU_NOM[int(m)]} {y}"
    except (ValueError, KeyError):
        return month
